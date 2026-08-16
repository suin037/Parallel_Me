"""KOWEPS 1~20차 Long Form → 변수사전·9영역 매핑·경량 종단패널.

원본 전체를 복제하지 않고 코드북과 Stata variable label을 함께 사용해 필요한 열만
선별한다. 산출물은 모두 재배포 금지 경로인 ``data/clean/koweps`` 아래에 둔다.

사용법::

    python preprocess/preprocess_koweps.py --audit-only
    python preprocess/preprocess_koweps.py
"""

from __future__ import annotations

import argparse
import json
import re
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path
from zipfile import ZipFile

import numpy as np
import pandas as pd
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
RAW = ROOT / "data" / "raw" / "koweps"
LONG = RAW / "long" / "koweps_hp01_20_long_260331.dta"
OUT = ROOT / "data" / "clean" / "koweps"

DOMAIN_KEYWORDS = {
    "career": ["취업", "직업", "직종", "종사상", "고용", "근로", "일자리", "실업"],
    "education": ["학력", "교육", "학교", "진학", "학업", "교육비"],
    "business": ["자영", "사업", "창업", "사업소득", "고용주"],
    "finance": ["소득", "임금", "자산", "재산", "부채", "지출", "생활비", "가처분"],
    "health": ["건강", "우울", "질병", "의료", "장애", "활동제약", "만성"],
    "housing": ["주거", "주택", "월세", "전세", "임대", "이사", "점유", "주거비"],
    "relationship": ["혼인", "배우자", "가족관계", "가구원", "자녀", "출산", "사회관계"],
    "lifestyle": ["생활만족", "여가", "근로시간", "수면", "시간사용", "삶의 만족"],
    "long_term_values": ["가치관", "미래", "만족도", "생활 만족", "복지인식", "계층"],
}

EVENT_KEYWORDS = {
    "employment_transition": ["경제활동상태", "종사상지위", "현재 취업상태"],
    "residential_move": ["지난 1년간 이사경험 여부", "이사횟수", "현 주택 거주기간"],
    "marriage_transition": ["혼인상태", "결혼상태", "배우자 유무"],
    "household_change": ["가구원수", "가구원 수", "출산 경험여부", "자녀수"],
}

OUTCOME_KEYWORDS = {
    "finance": ["가처분소득", "경상소득", "총생활비", "총부채", "총재산"],
    "health": ["주관적 건강", "우울", "건강상태", "의료비", "활동제약"],
    "housing": ["주거비", "주거환경", "주택유형", "점유형태", "주거 만족"],
    "relationship": ["가족관계 만족", "가족생활 만족", "사회적 관계"],
    "lifestyle": ["생활만족", "여가", "근로시간", "삶의 만족"],
}

CORE_COLUMNS = [
    "h_merkey", "h_pid", "year", "wv", "wv_num", "first_wv", "last_wv",
    "p_wsl_n_all", "p_wsc_n_all", "p_wgl_n_all", "p_wgc_n_all",
]


def _norm(value) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def find_codebook_zip() -> Path:
    candidates = sorted(RAW.glob("*.zip"), key=lambda p: p.stat().st_size)
    for path in candidates:
        with ZipFile(path) as zf:
            if any(name.lower().endswith(".xlsx") for name in zf.namelist()):
                return path
    raise FileNotFoundError("KOWEPS 결합데이터 코드북 ZIP을 찾지 못했습니다.")


def parse_codebooks(path: Path) -> pd.DataFrame:
    rows = []
    with ZipFile(path) as zf:
        for name in zf.namelist():
            if not name.lower().endswith(".xlsx"):
                continue
            wb = load_workbook(BytesIO(zf.read(name)), read_only=True, data_only=True)
            if "코딩북" not in wb.sheetnames:
                continue
            ws = wb["코딩북"]
            iterator = ws.iter_rows(values_only=True)
            header = [_norm(v) for v in next(iterator)]
            for values in iterator:
                item = dict(zip(header, values))
                variable = _norm(item.get(header[0]))
                if not variable:
                    continue
                variable = variable.splitlines()[0].strip()
                rows.append({
                    "variable": variable,
                    "waves": _norm(item.get(header[1])),
                    "section": _norm(item.get(header[2])),
                    "label": _norm(item.get(header[3])),
                    "question": _norm(item.get(header[4])),
                    "codebook_file": Path(name).name,
                })
    return pd.DataFrame(rows).drop_duplicates(["variable", "codebook_file"])


def stata_labels(path: Path) -> dict[str, str]:
    reader = pd.read_stata(path, iterator=True)
    return {str(k): _norm(v) for k, v in reader.variable_labels().items()}


def keyword_hits(text: str, keyword_map: dict[str, list[str]]) -> list[str]:
    return [key for key, words in keyword_map.items() if any(word in text for word in words)]


def event_hits(label: str) -> list[str]:
    """사건은 문항 본문의 우연한 단어가 아니라 변수설명(label)에서만 찾는다."""
    return keyword_hits(label, EVENT_KEYWORDS)


def build_dictionary(codebook: pd.DataFrame, labels: dict[str, str]) -> pd.DataFrame:
    cb = (codebook.sort_values("codebook_file")
          .drop_duplicates("variable", keep="last").set_index("variable"))
    rows = []
    for variable, stata_label in labels.items():
        item = cb.loc[variable].to_dict() if variable in cb.index else {}
        label = _norm(item.get("label")) or stata_label
        question = _norm(item.get("question"))
        haystack = f"{label} {question} {stata_label}"
        rows.append({
            "variable": variable, "label": label, "stata_label": stata_label,
            "question": question, "waves": _norm(item.get("waves")),
            "section": _norm(item.get("section")),
            "domains": "|".join(keyword_hits(haystack, DOMAIN_KEYWORDS)),
            "events": "|".join(event_hits(label)),
            "outcomes": "|".join(keyword_hits(haystack, OUTCOME_KEYWORDS)),
        })
    return pd.DataFrame(rows)


def selected_variables(dictionary: pd.DataFrame) -> list[str]:
    candidates = dictionary[
        dictionary["domains"].ne("") | dictionary["events"].ne("") |
        dictionary["outcomes"].ne("")
    ]["variable"].tolist()
    # 너무 넓은 단어 검색 때문에 수천 열을 다시 담지 않도록, 결과/사건 후보를 우선하고
    # 영역별로 앞쪽 35개까지만 감사용 패널에 포함한다.
    priority = dictionary[dictionary["events"].ne("") | dictionary["outcomes"].ne("")]["variable"].tolist()
    bounded = []
    for domain in DOMAIN_KEYWORDS:
        bounded.extend(dictionary[dictionary["domains"].str.split("|").apply(lambda x: domain in x)]["variable"].head(35))
    ordered = [*CORE_COLUMNS, *priority, *bounded, *candidates[:50]]
    return list(dict.fromkeys(v for v in ordered if v in set(dictionary.variable)))


def clean_missing(frame: pd.DataFrame) -> pd.DataFrame:
    for column in frame.select_dtypes(include=["number"]).columns:
        if column not in {"year", "wv", "wv_num", "first_wv", "last_wv"}:
            frame[column] = frame[column].mask(frame[column] < 0)
    return frame


def coverage(frame: pd.DataFrame, dictionary: pd.DataFrame) -> dict:
    by_var = dictionary.set_index("variable")
    variables = []
    for column in frame.columns:
        if column in CORE_COLUMNS:
            continue
        nonnull = int(frame[column].notna().sum())
        waves = int(frame.loc[frame[column].notna(), "wv"].nunique()) if nonnull else 0
        meta = by_var.loc[column] if column in by_var.index else {}
        variables.append({
            "variable": column, "label": meta.get("label", "") if hasattr(meta, "get") else "",
            "domains": meta.get("domains", "") if hasattr(meta, "get") else "",
            "events": meta.get("events", "") if hasattr(meta, "get") else "",
            "outcomes": meta.get("outcomes", "") if hasattr(meta, "get") else "",
            "non_null_n": nonnull, "coverage_rate": round(nonnull / len(frame), 4),
            "observed_waves": waves,
        })
    domain_summary = {}
    for domain in DOMAIN_KEYWORDS:
        items = [v for v in variables if domain in v["domains"].split("|")]
        domain_summary[domain] = {
            "candidate_variables": len(items),
            "usable_variables": sum(v["non_null_n"] >= 200 and v["observed_waves"] >= 2 for v in items),
            "top": sorted(items, key=lambda x: (x["observed_waves"], x["non_null_n"]), reverse=True)[:12],
        }
    return {
        "built_at": datetime.now(timezone.utc).isoformat(),
        "source": LONG.name, "rows": int(len(frame)),
        "people": int(frame["h_pid"].nunique()),
        "households": int(frame["h_merkey"].nunique()),
        "waves": sorted(int(x) for x in frame["wv"].dropna().unique()),
        "selected_columns": len(frame.columns), "domains": domain_summary,
        "variables": variables,
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--audit-only", action="store_true", help="코드북·DTA 변수사전만 생성")
    args = parser.parse_args()
    OUT.mkdir(parents=True, exist_ok=True)

    codebook_zip = find_codebook_zip()
    codebook = parse_codebooks(codebook_zip)
    labels = stata_labels(LONG)
    dictionary = build_dictionary(codebook, labels)
    dictionary.to_csv(OUT / "koweps_variable_dictionary.csv", index=False, encoding="utf-8-sig")

    mapping = {
        domain: dictionary[dictionary["domains"].str.split("|").apply(lambda x: domain in x)][
            ["variable", "label", "waves", "events", "outcomes"]
        ].to_dict("records")
        for domain in DOMAIN_KEYWORDS
    }
    (OUT / "koweps_domain_mapping.json").write_text(
        json.dumps(mapping, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(f"[dictionary] {len(dictionary):,}개 DTA 변수 / 코드북 {len(codebook):,}행")
    if args.audit_only:
        return

    usecols = selected_variables(dictionary)
    print(f"[panel] {len(usecols):,}개 열 선택, Long Form 읽는 중...")
    panel = pd.read_stata(LONG, columns=usecols, convert_categoricals=False)
    panel = clean_missing(panel).sort_values(["h_pid", "wv"]).reset_index(drop=True)
    panel.to_parquet(OUT / "koweps_life_panel.parquet", index=False)
    report = coverage(panel, dictionary)
    (OUT / "koweps_coverage_report.json").write_text(
        json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(f"[done] {len(panel):,}행 × {len(panel.columns):,}열 / {report['people']:,}명")


if __name__ == "__main__":
    main()
